import os
import sys
import json
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


GREEN_DARK = "1A5C2E"
GREEN_MID = "2D9E4F"
GREEN_LIGHT = "D6F5E3"
RED_DARK = "C0392B"
RED_LIGHT = "FDECEA"
AMBER_DARK = "B7770D"
AMBER_LIGHT = "FFF3CD"
GRAY_DARK = "2C2C2C"
GRAY_MID = "6B7280"
WHITE = "FFFFFF"


def style_header_cell(cell, text, bg=GREEN_DARK, fg=WHITE, bold=True, size=11):
    cell.value = text
    cell.font = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_data_cell(cell, value, bold=False, align="left", number_format=None):
    cell.value = value
    cell.font = Font(name="Arial", bold=bold, color=GRAY_DARK, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if number_format:
        cell.number_format = number_format


def status_fill(status):
    fills = {
        "CRITICAL": PatternFill("solid", fgColor=RED_LIGHT),
        "WARNING": PatternFill("solid", fgColor=AMBER_LIGHT),
        "OK": PatternFill("solid", fgColor=GREEN_LIGHT),
    }
    return fills.get(status, PatternFill("solid", fgColor=WHITE))


def status_font(status):
    colors = {
        "CRITICAL": RED_DARK,
        "WARNING": AMBER_DARK,
        "OK": GREEN_MID
    }
    return Font(name="Arial", bold=True, color=colors.get(status, GRAY_DARK), size=10)


def thin_border():
    side = Side(style="thin", color="D1D5DB")
    return Border(left=side, right=side, top=side, bottom=side)


def set_col_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def ensure_parent_dir(output_path):
    parent = os.path.dirname(os.path.abspath(output_path))
    if parent:
        os.makedirs(parent, exist_ok=True)


def parse_date(value, with_time=False):
    if not value:
        return ""
    try:
        if isinstance(value, datetime):
            return value.strftime("%d/%m/%Y %H:%M" if with_time else "%d/%m/%Y")
        s = str(value).replace("Z", "")
        dt = datetime.fromisoformat(s)
        return dt.strftime("%d/%m/%Y %H:%M" if with_time else "%d/%m/%Y")
    except Exception:
        return str(value)


def add_header_row(ws, row_num, ws_title, company, generated_by="Rivecor Eco Móvil 360"):
    ws.row_dimensions[row_num].height = 40
    ws.merge_cells(f"A{row_num}:O{row_num}")
    c = ws[f"A{row_num}"]
    c.value = f"RIVECOR ECO MÓVIL 360  ·  {ws_title.upper()}  ·  {company}"
    c.font = Font(name="Arial", bold=True, color=WHITE, size=13)
    c.fill = PatternFill("solid", fgColor=GREEN_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[row_num + 1].height = 18
    ws.merge_cells(f"A{row_num+1}:O{row_num+1}")
    c2 = ws[f"A{row_num+1}"]
    c2.value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  {generated_by}"
    c2.font = Font(name="Arial", color=GRAY_MID, size=9)
    c2.alignment = Alignment(horizontal="left", vertical="center")

    return row_num + 3


def build_tires_sheet(wb, data):
    ws = wb.create_sheet("Neumáticos por Equipo")
    ws.sheet_view.showGridLines = False

    row = add_header_row(ws, 1, "Neumáticos por Equipo", data.get("company", ""))

    headers = [
        "Equipo", "Código", "Posición", "Marca", "Medida",
        "Surco (mm)", "Presión (PSI)", "Km recorridos", "Km máx",
        "Vida útil %", "Estado"
    ]
    cols = list("ABCDEFGHIJK")
    widths = {
        "A": 22, "B": 14, "C": 26, "D": 14, "E": 16,
        "F": 14, "G": 16, "H": 16, "I": 12, "J": 14, "K": 14
    }
    set_col_widths(ws, widths)

    ws.row_dimensions[row].height = 28
    for i, h in enumerate(headers):
        style_header_cell(ws[f"{cols[i]}{row}"], h, bg=GREEN_MID)
    row += 1

    for eq in data.get("equipments", []):
      for tire in eq.get("tires", []):
        life_pct = None
        mileage = tire.get("mileage")
        max_mileage = tire.get("maxMileage")
        if mileage is not None and max_mileage not in (None, 0):
            life_pct = round((mileage / max_mileage) * 100, 1)

        vals = [
            eq.get("name", ""),
            eq.get("code", ""),
            tire.get("position", ""),
            tire.get("brand", ""),
            tire.get("size", ""),
            tire.get("currentDepth"),
            tire.get("pressure"),
            mileage,
            max_mileage,
            f"{life_pct}%" if life_pct is not None else "",
            tire.get("status", "OK"),
        ]

        status = tire.get("status", "OK")
        ws.row_dimensions[row].height = 22

        for i, v in enumerate(vals):
            c = ws[f"{cols[i]}{row}"]
            style_data_cell(c, v, align="center" if i > 2 else "left")
            c.border = thin_border()
            if status in ("CRITICAL", "WARNING"):
                c.fill = status_fill(status)
                if i == 10:
                    c.font = status_font(status)
            else:
                c.fill = PatternFill("solid", fgColor="F9FAFB")
                if i == 10:
                    c.font = status_font(status)
        row += 1

    ws.row_dimensions[row].height = 24
    ws.merge_cells(f"A{row}:E{row}")
    c = ws[f"A{row}"]
    total_tires = sum(len(eq.get("tires", [])) for eq in data.get("equipments", []))
    c.value = f"TOTAL: {total_tires} neumáticos"
    c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    c.fill = PatternFill("solid", fgColor=GREEN_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")

    for col in cols[5:]:
        ws[f"{col}{row}"].fill = PatternFill("solid", fgColor=GREEN_DARK)

    return ws


def build_history_sheet(wb, data):
    ws = wb.create_sheet("Historial Mantenciones")
    ws.sheet_view.showGridLines = False

    row = add_header_row(ws, 1, "Historial de Mantenciones", data.get("company", ""))

    headers = [
        "Fecha", "Equipo", "Mecánico", "Tipo",
        "Neumáticos trabajados", "Profundidad antes→después",
        "Presión antes→después", "Observaciones", "Próxima visita"
    ]
    cols = list("ABCDEFGHI")
    widths = {
        "A": 16, "B": 22, "C": 18, "D": 18, "E": 28,
        "F": 26, "G": 22, "H": 36, "I": 16
    }
    set_col_widths(ws, widths)

    ws.row_dimensions[row].height = 28
    for i, h in enumerate(headers):
        style_header_cell(ws[f"{cols[i]}{row}"], h, bg=GREEN_MID)
    row += 1

    type_map = {
        "INSPECTION": "Inspección",
        "ROTATION": "Rotación",
        "REPLACEMENT": "Reemplazo",
        "PRESSURE_CHECK": "Rev. Presión",
        "EMERGENCY": "Emergencia",
    }

    for maintenance in data.get("maintenances", []):
        tires_worked = maintenance.get("tiresWorked", [])

        tires_str = "; ".join([t.get("position", "") for t in tires_worked if t.get("position")])

        depth_str = "; ".join([
            f"{t.get('depthBefore', '?')}→{t.get('depthAfter', '?')} mm"
            for t in tires_worked
            if t.get("depthBefore") is not None or t.get("depthAfter") is not None
        ])

        pressure_str = "; ".join([
            f"{t.get('pressureBefore', '?')}→{t.get('pressureAfter', '?')} PSI"
            for t in tires_worked
            if t.get("pressureBefore") is not None or t.get("pressureAfter") is not None
        ])

        vals = [
            parse_date(maintenance.get("performedAt"), with_time=True),
            maintenance.get("equipmentName", ""),
            maintenance.get("mechanicName", ""),
            type_map.get(maintenance.get("type", ""), maintenance.get("type", "")),
            tires_str,
            depth_str,
            pressure_str,
            maintenance.get("observations", ""),
            parse_date(maintenance.get("nextScheduled")),
        ]

        fill = PatternFill("solid", fgColor="F9FAFB") if row % 2 == 0 else PatternFill("solid", fgColor=WHITE)
        ws.row_dimensions[row].height = 22

        for i, v in enumerate(vals):
            c = ws[f"{cols[i]}{row}"]
            style_data_cell(c, v, align="center" if i == 0 else "left")
            c.border = thin_border()
            c.fill = fill
        row += 1

    return ws


def build_cost_sheet(wb, data):
    ws = wb.create_sheet("Costo x Kilómetro")
    ws.sheet_view.showGridLines = False

    row = add_header_row(ws, 1, "Trazabilidad Costo x Kilómetro", data.get("company", ""))

    headers = [
        "Equipo", "Posición", "Marca", "Medida",
        "Precio compra ($)", "Costo mantención ($)", "Costo total ($)",
        "Km actuales", "Km máx",
        "Costo compra/km ($)", "Costo total/km ($)", "Vida útil %",
        "Km restantes", "Presión - pérdida %", "Recomendación"
    ]
    cols = [get_column_letter(i + 1) for i in range(len(headers))]
    widths = {
        "A": 22, "B": 22, "C": 14, "D": 16,
        "E": 18, "F": 18, "G": 16,
        "H": 14, "I": 12,
        "J": 18, "K": 18, "L": 14,
        "M": 16, "N": 20, "O": 36,
    }
    set_col_widths(ws, widths)

    ws.row_dimensions[row].height = 28
    for i, h in enumerate(headers):
        style_header_cell(ws[f"{cols[i]}{row}"], h, bg=GREEN_MID)
    row += 1

    first_data_row = row

    for item in data.get("costs", []):
        tire = item.get("tire", {})
        analysis = item.get("analysis", {})

        life_pct = analysis.get("lifeUsedPct", 0)

        if life_pct >= 90:
            row_fill = PatternFill("solid", fgColor=RED_LIGHT)
        elif life_pct >= 70:
            row_fill = PatternFill("solid", fgColor=AMBER_LIGHT)
        else:
            row_fill = PatternFill("solid", fgColor="F9FAFB" if row % 2 == 0 else WHITE)

        vals = [
            item.get("equipmentName", ""),
            tire.get("position", ""),
            tire.get("brand", ""),
            tire.get("size", ""),
            analysis.get("purchasePrice"),
            analysis.get("maintenanceCost", 0),
            analysis.get("totalCost"),
            analysis.get("currentKm"),
            analysis.get("maxKm"),
            analysis.get("costPerKm"),
            analysis.get("totalCostPerKm"),
            f"{life_pct}%",
            analysis.get("remainingKm"),
            f"{analysis.get('pressureLossPct', 0)}%",
            analysis.get("recommendation", ""),
        ]

        for i, v in enumerate(vals):
            c = ws[f"{cols[i]}{row}"]
            fmt = "$#,##0.0000" if i in {9, 10} else ("$#,##0" if i in {4, 5, 6} else None)
            style_data_cell(c, v, align="right" if i >= 4 else "left", number_format=fmt)
            c.border = thin_border()
            c.fill = row_fill

        row += 1

    if row > first_data_row:
        ws.row_dimensions[row].height = 26
        last_data_row = row - 1

        ws[f"A{row}"].value = "TOTALES"
        ws[f"A{row}"].font = Font(name="Arial", bold=True, color=WHITE, size=10)
        ws[f"A{row}"].fill = PatternFill("solid", fgColor=GREEN_DARK)
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")

        for col_letter in ["E", "F", "G", "H"]:
            c = ws[f"{col_letter}{row}"]
            c.value = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
            c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
            c.fill = PatternFill("solid", fgColor=GREEN_DARK)
            c.alignment = Alignment(horizontal="right", vertical="center")
            if col_letter in ("E", "F", "G"):
                c.number_format = "$#,##0"

        for col_letter in ["B", "C", "D", "I", "J", "K", "L", "M", "N", "O"]:
            ws[f"{col_letter}{row}"].fill = PatternFill("solid", fgColor=GREEN_DARK)

    return ws


def build_mechanics_sheet(wb, data):
    ws = wb.create_sheet("Mecánicos")
    ws.sheet_view.showGridLines = False

    row = add_header_row(ws, 1, "Ficha de Mecánicos", data.get("company", "Empresa"))

    headers = [
        "Nombre", "RUT", "Especialidad", "Teléfono",
        "Total mantenciones", "Neumáticos trabajados",
        "Reemplazos", "Rotaciones", "Inspecciones", "Última actividad"
    ]
    cols = list("ABCDEFGHIJ")
    widths = {
        "A": 24, "B": 14, "C": 20, "D": 16, "E": 20,
        "F": 22, "G": 14, "H": 14, "I": 14, "J": 20
    }
    set_col_widths(ws, widths)

    ws.row_dimensions[row].height = 28
    for i, h in enumerate(headers):
        style_header_cell(ws[f"{cols[i]}{row}"], h, bg=GREEN_MID)
    row += 1

    mechanics = data.get("mechanics", [])

    for mechanic in mechanics:
        stats = mechanic.get("stats", {})

        vals = [
            mechanic.get("name", ""),
            mechanic.get("rut", ""),
            mechanic.get("speciality", ""),
            mechanic.get("phone", ""),
            stats.get("totalMaintenances", 0),
            stats.get("totalTiresWorked", 0),
            stats.get("replacements", 0),
            stats.get("rotations", 0),
            stats.get("inspections", 0),
            parse_date(stats.get("lastActivity")),
        ]

        fill = PatternFill("solid", fgColor="F9FAFB") if row % 2 == 0 else PatternFill("solid", fgColor=WHITE)
        ws.row_dimensions[row].height = 22

        for i, v in enumerate(vals):
            c = ws[f"{cols[i]}{row}"]
            style_data_cell(c, v, align="center" if i >= 4 else "left")
            c.border = thin_border()
            c.fill = fill

        row += 1

    if mechanics:
        row += 2
        ws[f"A{row}"].value = "Mantenciones por mecánico"
        ws[f"A{row}"].font = Font(name="Arial", bold=True, color=GREEN_DARK, size=11)
        row += 1

        start_chart_row = row
        for mechanic in mechanics:
            ws[f"A{row}"].value = mechanic.get("name", "")
            ws[f"B{row}"].value = mechanic.get("stats", {}).get("totalMaintenances", 0)
            ws[f"C{row}"].value = mechanic.get("stats", {}).get("totalTiresWorked", 0)
            row += 1

        chart = BarChart()
        chart.type = "bar"
        chart.title = "Trabajo por mecánico"
        chart.y_axis.title = "Cantidad"
        chart.style = 10
        chart.width = 18
        chart.height = 12

        data_ref = Reference(ws, min_col=2, max_col=3, min_row=start_chart_row, max_row=row - 1)
        cats_ref = Reference(ws, min_col=1, min_row=start_chart_row, max_row=row - 1)
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, f"E{start_chart_row}")

    return ws


def generate_rivecor_excel(output_path, payload, sheet_filter=None):
    ensure_parent_dir(output_path)

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    company = payload.get("company", "Empresa")

    if sheet_filter in (None, "", "full", "tires"):
        build_tires_sheet(wb, {
            "company": company,
            "equipments": payload.get("equipments", [])
        })

    if sheet_filter in (None, "", "full", "history"):
        build_history_sheet(wb, {
            "company": company,
            "maintenances": payload.get("maintenances", [])
        })

    if sheet_filter in (None, "", "full", "costs"):
        build_cost_sheet(wb, {
            "company": company,
            "costs": payload.get("costs", [])
        })

    if sheet_filter in (None, "", "full", "mechanics"):
        build_mechanics_sheet(wb, {
            "company": company,
            "mechanics": payload.get("mechanics", [])
        })

    if not wb.sheetnames:
        ws = wb.create_sheet("Reporte")
        ws["A1"] = "Sin datos para generar reporte"

    wb.save(output_path)
    print(f"✅ Excel generado: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Uso: python3 generate_excel.py input.json output.xlsx [sheetFilter]")
        sys.exit(1)

    input_json = sys.argv[1]
    output_xlsx = sys.argv[2]
    sheet_filter = sys.argv[3] if len(sys.argv) >= 4 else None

    with open(input_json, "r", encoding="utf-8") as f:
        payload = json.load(f)

    generate_rivecor_excel(output_xlsx, payload, sheet_filter)
    sys.exit(0)